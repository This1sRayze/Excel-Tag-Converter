import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import subprocess
import os
import sys

class ExcelTagConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Tag Converter")
        self.root.geometry("800x700")
        self.root.configure(bg='#f0f0f0')
        
        self.input_file = None
        self.mapping_file = None
        self.data_type_mapping = {}
        
        # Color palette 
        self.area_colors = [
    "FF4472C4",
    "FF70AD47",
    "FFED7D31",
    "FF5B9BD5",
    "FFA5A5A5",
    "FFB38600",
    "FFB88A90",
    "FF8FB296",
    "FFB8A96F",
    "FF9AA4C4",
    "FF9FB19A",
    "FFBFA89E",
    "FF9FAFC6",
    "FFB07D5D",
    "FF6F94B8",
]
        
        self.setup_styles()
        self.setup_ui()
    
    def setup_styles(self):
        """Configure custom styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        bg_color = '#f0f0f0'
        frame_bg = '#ffffff'
        accent_color = '#0078d4'
        
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), background=bg_color, foreground='#2c3e50')
        style.configure('Header.TLabel', font=('Segoe UI', 10, 'bold'), background=frame_bg)
        style.configure('Info.TLabel', font=('Segoe UI', 9), background=frame_bg, foreground='#555555')
        style.configure('Card.TFrame', background=frame_bg, relief='flat', borderwidth=1)
        style.configure('Card.TLabelframe', background=frame_bg, relief='solid', borderwidth=1)
        style.configure('Card.TLabelframe.Label', font=('Segoe UI', 10, 'bold'), foreground='#2c3e50', background=frame_bg)
    
    def setup_ui(self):
        outer_container = tk.Frame(self.root, bg='#f0f0f0')
        outer_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        main_container = tk.Frame(outer_container, bg='#f0f0f0')
        main_container.pack(fill="both", expand=True, anchor="center")
        
        # Title
        title_label = ttk.Label(main_container, text="📊 Excel Tag Converter", style='Title.TLabel')
        title_label.pack(pady=(0, 15))
        
        columns_frame = tk.Frame(main_container, bg='#f0f0f0')
        columns_frame.pack(fill="x", expand=False, anchor="center")
        
        left_column = tk.Frame(columns_frame, bg='#f0f0f0')
        left_column.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        right_column = tk.Frame(columns_frame, bg='#f0f0f0')
        right_column.pack(side="left", fill="both", expand=True, padx=(10, 0))
        
        
        # File Selection Card
        file_frame = ttk.LabelFrame(left_column, text="📁 File Selection", style='Card.TLabelframe', padding=15)
        file_frame.pack(fill="x", pady=(0, 10))
        
        # Input file
        input_row = tk.Frame(file_frame, bg='#ffffff')
        input_row.pack(fill="x", pady=5)
        
        self.input_btn = tk.Button(input_row, text="Select Input Excel", command=self.select_input,
                                   bg='#0078d4', fg='white', font=('Segoe UI', 9, 'bold'),
                                   relief='flat', padx=15, pady=8, cursor='hand2', width=15)
        self.input_btn.pack(side="left", padx=(0, 10))
        
        self.input_label = ttk.Label(input_row, text="No input file", style='Info.TLabel')
        self.input_label.pack(side="left", fill="x")
        
        # Mapping file
        mapping_row = tk.Frame(file_frame, bg='#ffffff')
        mapping_row.pack(fill="x", pady=5)
        
        self.mapping_btn = tk.Button(mapping_row, text="Select Mapping Excel", command=self.select_mapping,
                                     bg='#5c6bc0', fg='white', font=('Segoe UI', 9, 'bold'),
                                     relief='flat', padx=15, pady=8, cursor='hand2', width=15)
        self.mapping_btn.pack(side="left", padx=(0, 10))
        
        self.mapping_label = ttk.Label(mapping_row, text="No mapping file", style='Info.TLabel')
        self.mapping_label.pack(side="left", fill="x")
        
        # Column Configuration Card
        col_frame = ttk.LabelFrame(left_column, text="⚙️ Column Configuration", style='Card.TLabelframe', padding=15)
        col_frame.pack(fill="x", pady=(0, 10))
        
        # Required columns
        req_label = ttk.Label(col_frame, text="Required Columns:", style='Header.TLabel')
        req_label.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))
        
        self.tag_name_col = self.create_column_input(col_frame, "Tag Name:", "Tag Name", 1)
        self.data_block_col = self.create_column_input(col_frame, "Data Block:", "Data Block", 2)
        self.desc_col = self.create_column_input(col_frame, "Description:", "Description", 3)
        self.type_col = self.create_column_input(col_frame, "UDT Type:", "UDT Type", 4)
        self.area_col = self.create_column_input(col_frame, "Area:", "Area", 5)
        
        # Optional columns
        opt_label = ttk.Label(col_frame, text="Optional Columns:", style='Header.TLabel')
        opt_label.grid(row=6, column=0, columnspan=2, sticky="w", pady=(15, 10))
        
        
        self.comments_col = self.create_column_input(col_frame, "Comments:", "Comments", 9)
        self.origin_col = self.create_column_input(col_frame, "Origin:", "Origin", 10)
        
        col_frame.columnconfigure(1, weight=1)
        
        # Info Card
        info_frame = ttk.LabelFrame(right_column, text="ℹ️ Mapping File Format", style='Card.TLabelframe', padding=15)
        info_frame.pack(fill="x", pady=(0, 10))
        
        info_text = ("Expected columns in mapping file:\n"
                    "  • UDT Type (e.g., ANL, BOOL, DIG_ALR)\n"
                    "  • Signal Type (e.g., HiAlarm, Status)\n"
                    "  • Data Type (REQUIRED - e.g., BOOL, REAL)\n\n"
                    "Array Types:\n"
                    "  Use the Start and End Columns\n"
                    "  Example: Array Start = 0, Array End = 16\n")
        
        info_label = ttk.Label(info_frame, text=info_text, style='Info.TLabel', justify="left")
        info_label.pack(anchor="w")
        
        # Log Card
        log_frame = ttk.LabelFrame(right_column, text="📝 Processing Log", style='Card.TLabelframe', padding=10)
        log_frame.pack(fill="x", expand=False)
        
        log_container = tk.Frame(log_frame, bg='#ffffff')
        log_container.pack(fill="both", expand=True)
        
        self.log_text = tk.Text(log_container, height=16, wrap="word", bg='#f8f9fa', 
                               fg='#2c3e50', font=('Consolas', 9), relief='flat', padx=10, pady=10)
        self.log_text.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(log_container, command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Process Button
        bottom_frame = tk.Frame(main_container, bg='#f0f0f0', height=70)
        bottom_frame.pack(fill="x", pady=(15, 0))
        bottom_frame.pack_propagate(False)
        
        self.process_btn = tk.Button(
            bottom_frame, 
            text="🚀 Process and Convert", 
            command=self.process_file,
            bg='#28a745', 
            fg='white', 
            font=('Segoe UI', 11, 'bold'),
            relief='flat', 
            padx=40, 
            pady=10, 
            cursor='hand2',
            activebackground='#218838',
            activeforeground='white',
            width=15
        )
        self.process_btn.pack(expand=True)
    
    def create_column_input(self, parent, label_text, default_value, row):
        """Helper to create column input fields"""
        label = ttk.Label(parent, text=label_text, style='Info.TLabel')
        label.grid(row=row, column=0, sticky="w", padx=5, pady=5)
        
        entry = ttk.Entry(parent, width=20, font=('Segoe UI', 9))
        entry.insert(0, default_value)
        entry.grid(row=row, column=1, padx=5, pady=5, sticky="ew")
        
        return entry
    
    def log(self, message):
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.root.update()
    
    def select_input(self):
        file = filedialog.askopenfilename(
            title="Select Input Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file:
            self.input_file = file
            self.input_label.config(text=f"✓ {Path(file).name}")
            self.log(f"✓ Selected input file: {Path(file).name}")
    
    def select_mapping(self):
        file = filedialog.askopenfilename(
            title="Select Data Type Mapping Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file:
            self.mapping_file = file
            self.mapping_label.config(text=f"✓ {Path(file).name}")
            self.log(f"✓ Selected mapping file: {Path(file).name}")
            self.load_mapping_file()
    
    def load_mapping_file(self):
        """Load the UDT type to signal type mapping"""
        try:
            df_mapping = pd.read_excel(self.mapping_file)
            
            required_cols = ['UDT Type', 'Signal Type']
            missing_cols = [col for col in required_cols if col not in df_mapping.columns]
            
            if missing_cols:
                raise ValueError(f"Mapping file must have 'UDT Type', 'Signal Type', and 'Data Type' columns.\nMissing: {', '.join(missing_cols)}")
            
            if 'Data Type' not in df_mapping.columns:
                raise ValueError("Mapping file must have 'Data Type' column")
            
            self.data_type_mapping = {}
            for _, row in df_mapping.iterrows():
                udt_type = str(row['UDT Type']).strip()
                signal_type = str(row['Signal Type']).strip()
                mapped_data_type = str(row['Data Type']).strip() if pd.notna(row['Data Type']) else ''
                
                array_info = self.parse_array_type(udt_type)
                base_type = array_info['base_type']
                
                if base_type not in self.data_type_mapping:
                    self.data_type_mapping[base_type] = {
                        'signals': [],
                        'is_array': array_info['is_array'],
                        'array_start': array_info['array_start'],
                        'array_end': array_info['array_end'],
                        'data_types': {}
                    }
                self.data_type_mapping[base_type]['signals'].append(signal_type)
                self.data_type_mapping[base_type]['data_types'][signal_type] = mapped_data_type
            
            self.log(f"✓ Loaded mapping for {len(self.data_type_mapping)} UDT types")
            for udt, info in self.data_type_mapping.items():
                array_str = f" [ARRAY {info['array_start']}..{info['array_end']}]" if info['is_array'] else ""
                self.log(f"  → {udt}{array_str}: {', '.join(info['signals'])}")
                
        except Exception as e:
            self.log(f"✗ Error loading mapping file: {str(e)}")
            messagebox.showerror("Error", f"Failed to load mapping file:\n{str(e)}")
            self.data_type_mapping = {}
    
    def parse_array_type(self, udt_type):
        """Parse array type like 'ARRAY[0..16] OF ANL' and return base type and array info"""
        array_match = re.match(r'ARRAY\[(\d+)\.\.(\d+)\]\s+OF\s+(.+)', udt_type, re.IGNORECASE)
        
        if array_match:
            return {
                'is_array': True,
                'array_start': int(array_match.group(1)),
                'array_end': int(array_match.group(2)),
                'base_type': array_match.group(3).strip()
            }
        else:
            return {
                'is_array': False,
                'array_start': None,
                'array_end': None,
                'base_type': udt_type
            }
    
    def extract_numeric_indices(self, scada_tag_path):
        """Extract all numeric indices from a scada tag path for proper numeric sorting"""
        indices = re.findall(r'\[(\d+)\]', str(scada_tag_path))
        return tuple(int(idx) for idx in indices) if indices else (0,)
    
    def extract_base_tag_path(self, scada_tag_path):
        """Extract base tag path without array indices"""
        return re.sub(r'\[\d+\]', '', str(scada_tag_path))

    def format_signal_label(self, signal_type):
        """Format signal type: replace underscores, insert spaces before uppercase, collapse spaces, uppercase."""
        if not signal_type:
            return ''
        s = str(signal_type)
        if s.replace('_', '').isupper():
            return s.replace('_', '').upper()

        s = s.replace('_', ' ')
        s = re.sub(r'(?<!^)(?=[A-Z])', ' ', s)
        s = ' '.join(s.split())
        s = s.upper()
        s = re.sub(r'\bHI\s+HI\b', 'HIHI', s)
        s = re.sub(r'\bLO\s+LO\b', 'LOLO', s)
        return s
    
    def get_signal_type_category(self, udt_type, description='', data_block='', area=''):
        """Determine signal type category (ANALOG, DIGITAL, COMM, CALCULATED) based on UDT type, description, data block, and area"""
        udt_lower = str(udt_type).lower().strip()
        desc_lower = str(description).lower().strip()
        data_block_lower = str(data_block).lower().strip()
        area_lower = str(area).lower().strip()
        
        # Check for CALCULATED type
        if 'position failure' in desc_lower or area_lower == 'diagnostics':
            return 'CALCULATED'
        
        # Check for COMM type
        comm_keywords_desc = ['deif', 'automaskin', 'mtu', 'consilium','nmea','modbus','gps']
        
        for keyword in comm_keywords_desc:
            if keyword in data_block_lower:
                return 'COMM'
            
        for keyword in comm_keywords_desc:
            if keyword in desc_lower:
                return 'COMM'
        comm_keywords_udt = ['deif', 'automaskin', 'mtu', 'consilium','nmea','modbus','gps']
        for keyword in comm_keywords_udt:
            if keyword in udt_lower:
                return 'COMM'
        
        # Check for ANALOG
        if udt_lower in ['anl', 'anl_tank'] or 'anl' in udt_lower:
            return 'ANALOG'
        
        # Check for DIGITAL
        digital_keywords = ['dig_alr', 'dig_alr_wo_inh', 'pump', 'bilge', 'valve', 'int']
        for keyword in digital_keywords:
            if keyword in udt_lower:
                return 'DIGITAL'
        
        return ''
    
    def format_worksheet(self, ws, is_scada=False, area_color=None, area_rows=None):
        """Apply professional formatting to worksheet"""
        header_fill = PatternFill(start_color='0078d4', end_color='0078d4', fill_type='solid')
        scada_header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        alt_row_fill = PatternFill(start_color='f0f8ff', end_color='f0f8ff', fill_type='solid')
        
        if area_color:
            header_fill = PatternFill(start_color=area_color, end_color=area_color, fill_type='solid')
        
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='Calibri', size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='d0d0d0'),
            right=Side(style='thin', color='d0d0d0'),
            top=Side(style='thin', color='d0d0d0'),
            bottom=Side(style='thin', color='d0d0d0')
        )
        
        for cell in ws[1]:
            cell.fill = scada_header_fill if is_scada else header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_fill = alt_row_fill
            if is_scada and area_rows:
                for area_name, area_info in area_rows.items():
                    if row_idx >= area_info['start'] and row_idx <= area_info['end']:
                        area_color_hex = area_info['color']
                        row_fill = PatternFill(start_color=area_color_hex, end_color=area_color_hex, fill_type='solid')
                        break
            elif row_idx % 2 == 0:
                row_fill = alt_row_fill
            
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                cell.fill = row_fill
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'
    
    def process_file(self):
        if not self.input_file:
            messagebox.showerror("Error", "Please select an input file")
            return
        
        # Ask user to select output file with custom filename
        default_filename = f"{Path(self.input_file).stem}_tagged.xlsx"
        output_file = filedialog.asksaveasfilename(
            title="Save Output File As",
            initialdir=str(Path(self.input_file).parent),
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_file:
            self.log("✗ Output file selection cancelled")
            return
        
        output_folder = str(Path(output_file).parent)
        
        try:
            self.log("\n" + "="*70)
            self.log("🚀 Starting processing...")
            self.process_btn.config(state="disabled")
            
            self.log(f"📖 Reading {Path(self.input_file).name}...")
            df = pd.read_excel(self.input_file)
            
            tag_name_col = self.tag_name_col.get()
            data_block_col = self.data_block_col.get()
            desc_col = self.desc_col.get()
            type_col = self.type_col.get()
            area_col = self.area_col.get()
            comments_col = self.comments_col.get()
            origin_col = self.origin_col.get()
            
            required_cols = [tag_name_col, data_block_col, desc_col, type_col, area_col]
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                raise ValueError(f"Required columns not found: {', '.join(missing_cols)}\nAvailable: {', '.join(df.columns)}")
            
            # Create empty columns for optional fields if they don't exist
            if comments_col not in df.columns:
                df[comments_col] = ''
                self.log(f"ℹ️ Created empty '{comments_col}' column")
            if origin_col not in df.columns:
                df[origin_col] = ''
                self.log(f"ℹ️ Created empty '{origin_col}' column")
            
            self.log(f"✓ Found {len(df)} rows to process")
            
            output_columns = []
            column_mapping = {}
            
            column_mapping['Data Block'] = data_block_col
            column_mapping['Tag Name'] = tag_name_col
            column_mapping['UDT Type'] = type_col
            column_mapping['Area'] = area_col
            column_mapping['Description'] = desc_col
            
            output_columns = ['Data Block', 'Tag Name', 'UDT Type', 'Area']

            if comments_col in df.columns:
                column_mapping['Comments'] = comments_col
                output_columns.append('Comments')
            if origin_col in df.columns:
                column_mapping['Origin'] = origin_col
                output_columns.append('Origin')

            output_columns.append('Description')
            comments_present = comments_col in df.columns
            origin_present = origin_col in df.columns
            
            df_output = pd.DataFrame()
            for col in output_columns:
                df_output[col] = df[column_mapping[col]]

            # Ensure exporter-friendly columns exist on area tabs
            if 'Origin' not in df_output.columns:
                df_output['Origin'] = df[origin_col] if origin_col in df.columns else ''
            if 'Signal Type' not in df_output.columns:
                df_output['Signal Type'] = ''
            if 'Is Alarm' not in df_output.columns:
                df_output['Is Alarm'] = False
            
            # Populate Signal Type for all rows
            for idx, row in df_output.iterrows():
                udt_type = row['UDT Type']
                description = row['Description']
                data_block = row['Data Block']
                area = row['Area']
                signal_type_category = self.get_signal_type_category(udt_type, description, data_block, area)
                df_output.at[idx, 'Signal Type'] = signal_type_category
            
            self.log(f"\n📝 Creating formatted Excel output...")
            output_file = Path(output_file)
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                areas = sorted(df[area_col].unique())
                
                for idx, area in enumerate(areas):
                    area_df = df_output[df_output['Area'] == area].copy()
                    
                    sheet_name = str(area)[:31]
                    for char in ['/', '\\', '*', '?', '[', ']', ':']:
                        sheet_name = sheet_name.replace(char, '_')
                    
                    # Enforce requested tagged columns order
                    desired_order = ['Data Block', 'Tag Name', 'UDT Type', 'Signal Type', 'Comments', 'Is Alarm', 'Origin', 'Description']
                    cols_present = [c for c in desired_order if c in area_df.columns]
                    if 'Description' in area_df.columns and 'Description' not in cols_present:
                        cols_present.append('Description')
                    if cols_present:
                        area_df = area_df[cols_present]

                        # Ensure area tab has exporter columns present and in a stable order
                        if 'Origin' not in area_df.columns:
                            area_df['Origin'] = df[origin_col] if origin_col in df.columns else ''
                        if 'Signal Type' not in area_df.columns:
                            area_df['Signal Type'] = ''
                        if 'Is Alarm' not in area_df.columns:
                            area_df['Is Alarm'] = False

                    area_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self.log(f"  ✓ Created '{sheet_name}' with {len(area_df)} rows")
                
                if self.data_type_mapping:
                    self.log("\n🔧 Generating SCADA SIGNAL tab...")
                    scada_rows = []
                    
                    for _, row in df.iterrows():
                        tag_name = row[tag_name_col] 
                        data_block = row[data_block_col]
                        description = row[desc_col]
                        area = row[area_col]
                        udt_type = str(row[type_col]).strip()
                        
                        
                        
                        array_type_match = re.match(r'ARRAY\[(\d+)\.\.(\d+)\]\s+OF\s+(.+)', udt_type, re.IGNORECASE)
                        
                        if array_type_match:
                            array_type_start = int(array_type_match.group(1))
                            array_type_end = int(array_type_match.group(2))
                            base_type = array_type_match.group(3).strip()

                            # Use the array bounds from the UDT definition
                            indices = list(range(array_type_start, array_type_end + 1))
                            
                            if base_type in self.data_type_mapping:
                                signal_types = self.data_type_mapping[base_type]['signals']
                                signal_type_category = self.get_signal_type_category(base_type, description, data_block, area)
                                
                                for signal_type in signal_types:
                                    for index in indices:
                                        mapped_data_type = self.data_type_mapping[base_type]['data_types'].get(signal_type, '')
                                        if mapped_data_type:
                                            scada_tag_path = f"{data_block}.{tag_name}[{index}].{signal_type}"
                                            data_type_col = mapped_data_type
                                        else:
                                            scada_tag_path = f"{data_block}.{tag_name}[{index}]"
                                            data_type_col = signal_type
                                        formatted_signal = self.format_signal_label(signal_type)
                                        sig_raw = str(signal_type) if signal_type is not None else ''
                                        is_data_type = bool(re.fullmatch(r'[A-Z0-9]+', sig_raw.replace('_', '')))
                                        if formatted_signal and not is_data_type and signal_type.upper() != 'STATUS':
                                            desc_with_signal = f"{description} {formatted_signal}".strip()
                                        else:
                                            desc_with_signal = description
                                        sig_str = str(signal_type) if signal_type is not None else ''
                                        suffix = str(scada_tag_path).split('.')[-1] if scada_tag_path else ''
                                        is_alarm = bool(re.search(r'(ALR|ALARM|HIHI|HI|LOLO|LO)', sig_str, re.IGNORECASE)) or bool(re.search(r'(ALR|ALARM|HIHI|HI|LOLO|LO)', str(suffix), re.IGNORECASE))
                                        scada_rows.append({
                                            'Area': area,
                                            'DB': data_block,
                                            'Scada Tag Path': scada_tag_path,
                                            'Type': signal_type,
                                            'Signal Type': signal_type_category,
                                            'Data Type': data_type_col,
                                            'Description': desc_with_signal,
                                            'Comments': (row[comments_col] if comments_present and pd.notna(row[comments_col]) else ''),
                                            'Origin': (row[origin_col] if origin_present and pd.notna(row[origin_col]) else ''),
                                            'Is Alarm': is_alarm
                                        })
                        
                        elif udt_type in self.data_type_mapping:
                            mapping_info = self.data_type_mapping[udt_type]
                            signal_types = mapping_info['signals']
                            signal_type_category = self.get_signal_type_category(udt_type, description, data_block, area)
                            
                            if mapping_info.get('is_array'):
                                indices = list(range(mapping_info['array_start'], mapping_info['array_end'] + 1))
                                
                                for signal_type in signal_types:
                                    for index in indices:
                                        mapped_data_type = self.data_type_mapping[udt_type]['data_types'].get(signal_type, '')
                                        if mapped_data_type:
                                            scada_tag_path = f"{data_block}.{tag_name}[{index}].{signal_type}"
                                            data_type_col = mapped_data_type
                                        else:
                                            scada_tag_path = f"{data_block}.{tag_name}[{index}]"
                                            data_type_col = signal_type
                                        formatted_signal = self.format_signal_label(signal_type)
                                        sig_raw = str(signal_type) if signal_type is not None else ''
                                        is_data_type = bool(re.fullmatch(r'[A-Z0-9]+', sig_raw.replace('_', '')))
                                        if formatted_signal and not is_data_type and signal_type.upper() != 'STATUS':
                                            desc_with_signal = f"{description} {formatted_signal}".strip()
                                        else:
                                            desc_with_signal = description
                                        sig_str = str(signal_type) if signal_type is not None else ''
                                        suffix = str(scada_tag_path).split('.')[-1] if scada_tag_path else ''
                                        is_alarm = bool(re.search(r'(ALR|ALARM|HIHI|HI|LOLO|LO)', sig_str, re.IGNORECASE)) or bool(re.search(r'(ALR|ALARM|HIHI|HI|LOLO|LO)', str(suffix), re.IGNORECASE))
                                        scada_rows.append({
                                            'Area': area,
                                            'DB': data_block,
                                            'Scada Tag Path': scada_tag_path,
                                            'Type': signal_type,
                                            'Signal Type': signal_type_category,
                                            'Data Type': data_type_col,
                                            'Description': desc_with_signal,
                                            'Comments': (row[comments_col] if comments_present and pd.notna(row[comments_col]) else ''),
                                            'Origin': (row[origin_col] if origin_present and pd.notna(row[origin_col]) else ''),
                                            'Is Alarm': is_alarm
                                        })
                            else:
                                for signal_type in signal_types:
                                    mapped_data_type = self.data_type_mapping[udt_type]['data_types'].get(signal_type, '')
                                    if mapped_data_type:
                                        scada_tag_path = f"{data_block}.{tag_name}.{signal_type}"
                                        data_type_col = mapped_data_type
                                    else:
                                        scada_tag_path = f"{data_block}.{tag_name}"
                                        data_type_col = signal_type
                                    formatted_signal = self.format_signal_label(signal_type)
                                    sig_raw = str(signal_type) if signal_type is not None else ''
                                    is_data_type = bool(re.fullmatch(r'[A-Z0-9]+', sig_raw.replace('_', '')))
                                    if formatted_signal and not is_data_type and signal_type.upper() != 'STATUS':
                                        desc_with_signal = f"{description} {formatted_signal}".strip()
                                    else:
                                        desc_with_signal = description
                                    sig_str = str(signal_type) if signal_type is not None else ''
                                    suffix = str(scada_tag_path).split('.')[-1] if scada_tag_path else ''
                                    is_alarm = bool(re.search(r'(ALR|ALARM|HIHI|HI|LOLO|LO)', sig_str, re.IGNORECASE)) or bool(re.search(r'(ALR|ALARM|HIHI|HI|LOLO|LO)', str(suffix), re.IGNORECASE))
                                    scada_rows.append({
                                        'Area': area,
                                        'DB': data_block,
                                        'Scada Tag Path': scada_tag_path,
                                        'Type': signal_type,
                                        'Signal Type': signal_type_category,
                                        'Data Type': data_type_col,
                                        'Description': desc_with_signal,
                                        'Comments': (row[comments_col] if comments_present and pd.notna(row[comments_col]) else ''),
                                        'Origin': (row[origin_col] if origin_present and pd.notna(row[origin_col]) else ''),
                                        'Is Alarm': is_alarm
                                    })
                        
                        
                    
                    if scada_rows:
                        scada_df = pd.DataFrame(scada_rows)
                        scada_df['Area'] = scada_df['Area'].astype(str).str.strip()
                        scada_df['Scada Tag Path'] = scada_df['Scada Tag Path'].astype(str).str.strip()
                        try:
                            areas_list = [str(a) for a in areas]
                            scada_df['Area'] = pd.Categorical(scada_df['Area'], categories=areas_list, ordered=True)
                        except Exception:
                            pass

                        # Sort with numeric-aware index handling for array indices
                        scada_df['_numeric_indices'] = scada_df['Scada Tag Path'].apply(self.extract_numeric_indices)
                        scada_df['_base_tag'] = scada_df['Scada Tag Path'].apply(self.extract_base_tag_path)
                        scada_df = scada_df.sort_values(by=['Area', '_base_tag', '_numeric_indices']).reset_index(drop=True)
                        scada_df = scada_df.drop(['_numeric_indices', '_base_tag'], axis=1)

                        # Compute area row map BEFORE dropping Area (used for coloring)
                        area_row_map = {}
                        current_row = 2
                        area_color_idx = 0
                        for area_name in scada_df['Area'].unique():
                            area_count = len(scada_df[scada_df['Area'] == area_name])
                            area_row_map[area_name] = {
                                'start': current_row,
                                'end': current_row + area_count - 1,
                                'color': self.area_colors[area_color_idx % len(self.area_colors)]
                            }
                            current_row += area_count
                            area_color_idx += 1

                        # Final SCADA sheet: DB, Scada Tag Path, Type, Signal Type, Data Type, Comments, Origin, Description
                        final_cols = ['DB', 'Scada Tag Path', 'Type', 'Signal Type', 'Data Type', 'Comments', 'Origin', 'Description']
                        final_cols = [c for c in final_cols if c in scada_df.columns]
                        final_scada = scada_df[final_cols]
                        final_scada.to_excel(writer, sheet_name='SCADA_SIGNAL', index=False)
                        self.log(f"  ✓ Created SCADA_SIGNAL with {len(final_scada)} rows")
            
            self.log("\n🎨 Applying professional formatting...")
            wb = load_workbook(output_file)
            
            area_idx = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                is_scada = sheet_name == 'SCADA_SIGNAL'
                
                area_color = None
                area_rows_info = None
                
                if not is_scada:
                    area_color = self.area_colors[area_idx % len(self.area_colors)]
                    area_idx += 1
                    self.format_worksheet(ws, is_scada, area_color)
                else:
                    self.format_worksheet(ws, is_scada, None, area_row_map)
                
                self.log(f"  ✓ Formatted '{sheet_name}'")
            
            wb.save(output_file)
            
            self.log(f"\n✅ SUCCESS! Output saved to:")
            self.log(f"   {output_file}")
            self.log("="*70)
            messagebox.showinfo("Success", 
                              f"Processing complete! ✓\n\n"
                              f"Processed: {len(df)} rows\n"
                              f"Created: {len(areas)} area tabs\n"
                              f"Output: {output_file.name}")
            
            # Open the output folder
            if sys.platform == 'win32':
                os.startfile(output_folder)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', output_folder])
            else:
                subprocess.Popen(['xdg-open', output_folder])
            
        except Exception as e:
            self.log(f"\n❌ ERROR: {str(e)}")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        
        finally:
            self.process_btn.config(state="normal")

    
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelTagConverter(root)
    root.mainloop()